from random import choice

import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime

from linked_list import LinkedList, DoAn
from file_handler import FileHandler
from gui import AppGUI


class Controller:
    def __init__(self):
        self.linked_list = LinkedList()
        
        print("=== ĐANG KHỞI ĐỘNG CHƯƠNG TRÌNH QUẢN LÝ ĐỒ ÁN ===")
        FileHandler.load(self.linked_list)
        
        self.gui = AppGUI(self)
        
        total = len(self.get_all())
        print(f" Khởi động thành công! Tổng số đồ án: {total}")
        if total == 0:
            print(" (Chưa có dữ liệu nào. Bạn có thể thêm mới hoặc nhập từ Excel)")
    def get_all(self):
        return self.linked_list.to_list()

    def them_moi(self):
        self.show_modal("Thêm đồ án mới")

    def sua(self):
        selected = self.gui.tree.selection()
        if not selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn đồ án cần sửa!")
            return
        
        item = selected[0]
        values = self.gui.tree.item(item)["values"]
        ma = values[1]                               
        
        doan = self.linked_list.tim_theo_ma(ma)
        if doan:
            self.show_modal("Sửa đồ án", doan)
        else:
            messagebox.showerror("Lỗi", "Không tìm thấy đồ án!")

    def xoa(self):
        selected = self.gui.tree.selection()
        if not selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn đồ án để xóa!")
            return
        if messagebox.askyesno("Xác nhận", f"Xóa {len(selected)} đồ án đã chọn?"):
            for item in selected:
                values = self.gui.tree.item(item)["values"]
                ma = values[1]                       # Mã đồ án ở index 1
                self.linked_list.xoa(ma)
            FileHandler.save(self.linked_list)
            self.gui.refresh_table()

    def tim_kiem(self, event=None):
        keyword = self.gui.search_entry.get().strip().lower()
        if not keyword:
            self.gui.refresh_table()
            return
        filtered = [da for da in self.get_all() 
                    if keyword in da.ma.lower() or 
                       keyword in da.ten.lower() or 
                       keyword in da.hoten.lower() or
                       keyword in da.gvhd.lower() or
                       keyword in da.lop.lower() or 
                       keyword in da.namhoc.lower() or      
                       keyword in da.trangthai.lower() or  
                       keyword in str(da.diem).lower()]
        self.gui.refresh_table(filtered)

    def sap_xep(self, choice):
        """Sắp xếp theo lựa chọn"""
        sort_map = {
            "Mã đề tài": self.linked_list.sap_xep_merge_theo_ma,
            "MSSV": self.linked_list.sap_xep_merge_theo_mssv,
            "Tên đề tài": self.linked_list.sap_xep_merge_theo_ten,
            "Họ tên SV": self.linked_list.sap_xep_merge_theo_hoten,
            "Lớp": self.linked_list.sap_xep_merge_theo_lop,
            "GVHD": self.linked_list.sap_xep_merge_theo_gvhd,
            "Điểm cao → thấp": lambda: self.linked_list.sap_xep_merge_theo_diem(reverse=True),
            "Điểm thấp → cao": lambda: self.linked_list.sap_xep_merge_theo_diem(reverse=False),
            "Trạng thái": self.linked_list.sap_xep_theo_trangthai,
        }
        if choice in sort_map:
            sort_map[choice]()
            self.gui.refresh_table()

    def thong_ke(self):
        lst = self.get_all()

        tong = len(lst)
        
        dang_thuc_hien = sum(1 for x in lst if x.trangthai == "Đang thực hiện")
        da_bao_cao     = sum(1 for x in lst if x.trangthai == "Đã báo cáo")
        da_hoan_thanh  = sum(1 for x in lst if x.trangthai == "Đã hoàn thành")

        diem_list = [x.diem for x in lst if x.trangthai == "Đã báo cáo" and x.diem > 0]
        diem_tb = sum(diem_list) / len(diem_list) if diem_list else 0.0

    def xuat_excel(self):
        try:
            from openpyxl import Workbook
            file_path = filedialog.asksaveasfilename(
                initialfile=f"DoAn_{datetime.now().strftime('%Y%m%d')}",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            headers = ["STT", "Mã Đề Tài", "Tên Đề Tài", "MSSV", "Họ tên SV", "GVHD", "Năm học", "Trạng thái", "Điểm"]
            for col, h in enumerate(headers, 1):
                ws.cell(1, col, h)

            for r, da in enumerate(self.get_all(), 2):
                ws.cell(r, 1, r-1)
                ws.cell(r, 2, da.ma)
                ws.cell(r, 3, da.ten)
                ws.cell(r, 4, da.mssv)
                ws.cell(r, 5, da.hoten)
                ws.cell(r, 6, da.gvhd)
                ws.cell(r, 7, da.namhoc)
                ws.cell(r, 8, da.trangthai)
                ws.cell(r, 9, da.diem)

            wb.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất file Excel:\n{file_path}")
        except ImportError:
            messagebox.showerror("Lỗi", "Chưa cài thư viện openpyxl.\nChạy lệnh:\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xuất Excel:\n{str(e)}")

    def nhap_excel(self):
        """Nhập dữ liệu từ file Excel"""
        try:
            file_path = filedialog.askopenfilename(
                title="Chọn file Excel",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not file_path:
                return

            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            count = 0
            skipped = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 2 or not str(row[1] if len(row) > 1 else "").strip():
                    continue

                try:
                    ma = str(row[1]).strip()
                    ten = str(row[2]).strip() if len(row) > 2 else ""
                    mssv = str(row[3]).strip() if len(row) > 3 else ""
                    hoten = str(row[4]).strip() if len(row) > 4 else ""
                    lop = str(row[5]).strip() if len(row) > 5 else ""
                    gvhd = str(row[6]).strip() if len(row) > 6 else ""
                    namhoc = str(row[7]).strip() if len(row) > 7 else "2025-2026"
                    trangthai = str(row[8]).strip() if len(row) > 8 else "Đang thực hiện"

                    diem = 0.0
                    if trangthai == "Đã báo cáo" and len(row) > 9:
                        try:
                            diem = float(row[9])
                            if not 0 <= diem <= 10:
                                diem = 0.0
                        except:
                            diem = 0.0

                    doan = DoAn(ma, ten, mssv, hoten, gvhd, lop, namhoc, trangthai, diem)

                    if self.linked_list.them(doan):
                        count += 1
                    else:
                        skipped += 1
                except:
                    skipped += 1

            FileHandler.save(self.linked_list)
            self.gui.refresh_table()

            messagebox.showinfo("Nhập Excel", 
                f"Đã nhập thành công {count} đồ án!\nBỏ qua {skipped} dòng (trùng hoặc lỗi).")

        except ImportError:
            messagebox.showerror("Lỗi", "Chưa cài openpyxl.\nChạy: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi nhập Excel:\n{str(e)}")
    def show_modal(self, title, doan=None):
        """Hiển thị form thêm / sửa đồ án"""
        modal = ctk.CTkToplevel(self.gui)
        modal.title(title)
        modal.geometry("720x700")
        modal.grab_set()

        entries = {}
        fields = ["Mã Đề Tài", "Tên Đề Tài", "MSSV", "Họ tên SV", "Lớp", "GVHD", 
                  "Năm học", "Trạng thái", "Điểm"]

        # Giá trị mặc định
        if doan:
            defaults = [doan.ma, doan.ten, doan.mssv, doan.hoten, doan.lop, 
                        doan.gvhd, doan.namhoc, doan.trangthai, f"{doan.diem:.1f}"]
        else:
            defaults = ["", "", "", "", "", "", "2025-2026", "Đang thực hiện", "0.0"]

        for i, field in enumerate(fields):
            frame = ctk.CTkFrame(modal)
            frame.pack(fill="x", padx=40, pady=8)

            ctk.CTkLabel(frame, text=field + ":", width=140, anchor="w",
                        font=ctk.CTkFont(size=14, weight="bold")).pack(side="left")

            if field == "Trạng thái":
                combo = ctk.CTkComboBox(frame, 
                                      values=["Đang thực hiện", "Đã báo cáo", "Đã hoàn thành"],
                                      width=360, height=36)
                combo.set(defaults[i])
                combo.pack(side="left", fill="x", expand=True)
                entries[field] = combo
                combo.configure(command=lambda v: self.toggle_diem_entry(entries))

            elif field == "Điểm":
                self.diem_entry = ctk.CTkEntry(frame, width=360, font=ctk.CTkFont(size=13))
                self.diem_entry.insert(0, defaults[i])
                self.diem_entry.pack(side="left", fill="x", expand=True)
                entries[field] = self.diem_entry
            else:
                entry = ctk.CTkEntry(frame, width=360, font=ctk.CTkFont(size=13))
                entry.insert(0, defaults[i])
                entry.pack(side="left", fill="x", expand=True)
                entries[field] = entry

        # Nút Lưu
        ctk.CTkButton(modal, text="💾 Lưu", width=180, height=50,
                      font=ctk.CTkFont(size=16, weight="bold"),
                      command=lambda: self._save_from_modal(modal, entries, doan)
                      ).pack(pady=25)

        self.toggle_diem_entry(entries)

    def _save_from_modal(self, modal, entries, doan=None):
        """Xử lý logic lưu dữ liệu từ modal"""
        try:
            trang_thai = entries["Trạng thái"].get().strip()
            diem_str = entries["Điểm"].get().strip()

            if trang_thai == "Đã báo cáo":
                try:
                    diem = float(diem_str) if diem_str else 0.0
                    if not 0 <= diem <= 10:
                        raise ValueError("Điểm phải từ 0.0 đến 10.0")
                except:
                    raise ValueError("Điểm phải là số hợp lệ trong khoảng 0.0 - 10.0!")
            else:
                diem = 0.0

            doan_moi = DoAn(
                entries["Mã Đề Tài"].get().strip(),
                entries["Tên Đề Tài"].get().strip(),
                entries["MSSV"].get().strip(),
                entries["Họ tên SV"].get().strip(),
                entries["GVHD"].get().strip(),
                entries["Lớp"].get().strip(),
                entries["Năm học"].get().strip(),
                trang_thai,
                diem
            )

            if doan is None:  # Thêm mới
                if not doan_moi.ma:
                    raise ValueError("Mã đề tài không được để trống!")
                if self.linked_list.them(doan_moi):
                    FileHandler.save(self.linked_list)
                    self.gui.refresh_table()
                    modal.destroy()
                    messagebox.showinfo("Thành công", "Thêm đề tài mới thành công!")
                else:
                    messagebox.showerror("Lỗi", "Mã đề tài hoặc MSSV đã tồn tại!")
            else:  # Sửa
                if self.linked_list.sua(doan.ma, doan_moi):
                    FileHandler.save(self.linked_list)
                    self.gui.refresh_table()
                    modal.destroy()
                    messagebox.showinfo("Thành công", "Cập nhật thành công!")
                else:
                    messagebox.showerror("Lỗi", "Không tìm thấy đề tài để sửa!")
                    
        except ValueError as ve:
            messagebox.showerror("Lỗi nhập liệu", str(ve))
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {str(e)}")

    def toggle_diem_entry(self, entries):
        """Bật/tắt ô nhập điểm theo trạng thái"""
        trang_thai = entries["Trạng thái"].get()
        diem_entry = entries.get("Điểm")
        if not diem_entry:
            return

        if trang_thai == "Đã báo cáo":
            diem_entry.configure(state="normal")
            if diem_entry.get().strip() in ["0.0", "0", ""]:
                diem_entry.delete(0, "end")
                diem_entry.focus()
        else:
            diem_entry.configure(state="disabled")
            diem_entry.delete(0, "end")
            diem_entry.insert(0, "0.0")

if __name__ == "__main__":
    ctk.set_appearance_mode("light")
    controller = Controller()
    controller.gui.mainloop()